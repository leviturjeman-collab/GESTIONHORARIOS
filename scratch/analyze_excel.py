import openpyxl
from openpyxl.utils import get_column_letter
import sys

filepath = r"c:\Users\Levi\GestionHorariosLEVI\HorariosFuengirolaClaude.xlsx"
output_file = r"c:\Users\Levi\GestionHorariosLEVI\scratch\excel_analysis_output.txt"

wb = openpyxl.load_workbook(filepath, data_only=True)

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("=" * 80 + "\n")
    f.write("ANALISIS COMPLETO DEL EXCEL DE CUADRANTES\n")
    f.write("=" * 80 + "\n")
    
    f.write(f"\nHOJAS DEL LIBRO: {wb.sheetnames}\n")
    f.write(f"Total hojas: {len(wb.sheetnames)}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"\n{'=' * 80}\n")
        f.write(f"HOJA: '{sheet_name}'\n")
        f.write(f"   Dimensiones: {ws.dimensions}\n")
        f.write(f"   Filas usadas: {ws.max_row}\n")
        f.write(f"   Columnas usadas: {ws.max_column}\n")
        f.write(f"{'=' * 80}\n")
        
        # Merged cells
        if ws.merged_cells.ranges:
            f.write(f"\n   CELDAS COMBINADAS ({len(ws.merged_cells.ranges)}):\n")
            for merged in ws.merged_cells.ranges:
                f.write(f"      {merged}\n")
        
        # Dump ALL cell data
        f.write(f"\n   CONTENIDO CELDA POR CELDA:\n")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell_info = f"      [{cell.coordinate}] = {repr(cell.value)}"
                    cell_info += f"  (tipo: {type(cell.value).__name__})"
                    
                    # Font
                    if cell.font:
                        font_details = []
                        if cell.font.bold:
                            font_details.append("BOLD")
                        if cell.font.italic:
                            font_details.append("italic")
                        if cell.font.name:
                            font_details.append(f"font={cell.font.name}")
                        if cell.font.size:
                            font_details.append(f"size={cell.font.size}")
                        if cell.font.color and cell.font.color.rgb and str(cell.font.color.rgb) != '00000000':
                            font_details.append(f"color={cell.font.color.rgb}")
                        if font_details:
                            cell_info += f"  [{'|'.join(font_details)}]"
                    
                    # Fill
                    try:
                        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                            rgb = str(cell.fill.start_color.rgb)
                            if rgb and rgb != '00000000':
                                cell_info += f"  [bg={rgb}]"
                    except:
                        pass
                    
                    # Alignment
                    if cell.alignment:
                        align_details = []
                        if cell.alignment.horizontal:
                            align_details.append(f"h={cell.alignment.horizontal}")
                        if cell.alignment.vertical:
                            align_details.append(f"v={cell.alignment.vertical}")
                        if cell.alignment.wrap_text:
                            align_details.append("wrap")
                        if align_details:
                            cell_info += f"  [align: {','.join(align_details)}]"
                    
                    # Number format
                    if cell.number_format and cell.number_format != 'General':
                        cell_info += f"  [fmt={cell.number_format}]"
                    
                    f.write(cell_info + "\n")
        
        # Column summary
        f.write(f"\n   RESUMEN POR COLUMNAS:\n")
        for col_idx in range(1, min(ws.max_column + 1, 50)):
            values = []
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    values.append((row_idx, val))
            if values:
                col_letter = get_column_letter(col_idx)
                unique_vals = set(str(v[1]) for v in values)
                f.write(f"\n   Columna {col_letter}: {len(values)} celdas con datos, {len(unique_vals)} valores unicos\n")
                for v in sorted(unique_vals):
                    f.write(f"     - {v}\n")

print(f"Output written to {output_file}")
